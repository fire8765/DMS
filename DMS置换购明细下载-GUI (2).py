#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import numpy as np
import pandas as pd
import openpyxl
import xlrd
import datetime
import time
import tkinter as tk # 导入tkinter模块
from tkinter import filedialog
from tkinter.simpledialog import askinteger, askfloat, askstring
import tkinter.messagebox as msgbox
import DMS_dwzh as dw
# 定义GUI：我的函数库
          
class MyGUI:
    def __init__(self):
        self.endDate=datetime.datetime.now().strftime('%Y-%m-%d')
        self.startDate=datetime.datetime.now().strftime('%Y-%m-%d')
        self.Folderpath=os.getcwd()
        self.skcode='17061001'
        self.customer_name=''
        window = tk.Tk() # 生成一个窗口对象
#         window.geometry("440x120") 
        window.title('DMS置换购明细下载工具') # 设置窗口标题
#         window.iconphoto(True, tk.PhotoImage(file='favicon.ico'))
        # frame1：市场监控
        frame0 = tk.Frame(window) # 生成第一个框架
        frame0.pack() # 将框架1放入窗口
        tk.Label(frame0, text = '1、选择日期和保存路径-可重复选择',).pack(side='left') # 生成框架标题
        tk.Button(frame0, text="开始日期",command =self.startdt).pack(side='left') # 插入按钮
        tk.Button(frame0, text="结束日期",command =self.enddt).pack(side='left') # 插入按钮
        tk.Button(frame0, text="保存路径",command =self.savefile).pack(side='left') # 插入按钮
        
        frame01 = tk.Frame(window) # 生成第一个框架
        frame01.pack() # 将框架1放入窗口   
#         tk.Label(frame01, text = '2、请选择批量或者单店-可重复选择',).pack(side='left') # 生成框架标题
#         tk.Button(frame01, text="批量下载",command =self.piliang).pack(side='left') # 插入按钮
#         tk.Button(frame01, text="单店下载",command =self.dandian).pack(side='left') # 插入按钮        
        
        
        # frame1：市场监控
#         frame1 = tk.Frame(window) # 生成第一个框架
#         frame1.pack() # 将框架1放入窗口
#         tk.Label(frame1, text = '库存发运类',).pack(side='left') # 生成框架标题
#         tk.Button(frame1, text="库存明细",command =self.kucundw).pack(side='left') # 插入按钮
#         tk.Button(frame1, text="发运明细",command =self.fayun_data).pack(side='left') # 插入按钮
#         tk.Button(frame1, text="订单绑定明细",command =self.dd_dw).pack(side='left') # 插入按钮
#         tk.Button(frame1, text="发运指令明细",command =self.fayunzl_dw).pack(side='left') # 插入按钮
#         tk.Button(frame1, text="发运资源明细",command =self.flzymx_dw).pack(side='left') # 插入按钮

#         frame2：期货数据库管理
        frame2 = tk.Frame(window) # 生成第二个框架
        frame2.pack() # 将框架2放入窗口
        tk.Label(frame2, text = '置换明细下载').pack(side='left') # 生成框架标题
        tk.Button(frame2, text="批量下载置换明细",command =self.zhihuan_down).pack(side='left') # 插入按钮
        tk.Button(frame2, text="单店下载置换明细",command =self.dandian).pack(side='left') # 插入按钮
        tk.Button(frame2, text="单独下载置换明细",command =self.zhihuan_down1).pack(side='left') # 插入按钮

        frame3 = tk.Frame(window) # 生成第二个框架
        frame3.pack() # 将框架2放入窗口
        tk.Label(frame3, text = '客户明细处理').pack(side='left') # 生成框架标题
        tk.Button(frame3, text="本品旧车明细",command =self.bpsx).pack(side='left') # 插入按钮
        tk.Button(frame3, text="他牌旧车明细",command =self.tpsx).pack(side='left') # 插入按钮
        tk.Button(frame3, text="筛选销售明细",command =self.mxsx).pack(side='left') # 插入按钮        
        # frame3：杂放工具
#         frame3 = tk.Frame(window) # 生成第三个框架
#         frame3.pack() # 将框架3放入窗口
#         tk.Label(frame3, text = '其他下载').pack(side='left') # 生成框架标题
#         tk.Button(frame3, text="折让明细",command =self.zherang_dw).pack(side='left') # 插入按钮
#         tk.Button(frame3, text="财务可用余额",command =self.cwye_dw).pack(side='left') # 插入按钮
#         tk.Button(frame3, text="车辆验证查询",command =dw.test).pack(side='left') # 插入按钮
       
        window.mainloop() # 创建事件循环（不必理解，照抄即可）
    def savefile(self):
#         root = tk.Tk()
#         self.window.withdraw()
#         选择文件夹
        self.Folderpath = filedialog.askdirectory()
        # 选择文件
#         Filepath = filedialog.askopenfilename()
#         打印文件夹路径
        print('Folderpath:', self.Folderpath)
#         # 打印文件路径
#         print('Filepath:', Filepath)
        return self.Folderpath
    def getfile(self):
#         root = tk.Tk()
#         self.window.withdraw()
        # 选择文件夹
#         Folderpath = filedialog.askdirectory()
        # 选择文件
        Filepath = filedialog.askopenfilename()
        # 打印文件夹路径
#         print('Folderpath:', Folderpath)
        # 打印文件路径
        print('Filepath:', Filepath)
        return Filepath
    def fmdate(self,date):
        try:
            time.strptime(date,"%Y-%m-%d")
            return True
        except:
            return False
    def piliang(self):
        try:
            data=r'信息表4.xlsx'
            df = pd.read_excel(data,header=None,index_col=3)
            self.username = df[0].tolist() # 此处改用户名
            self.password1 = df[1].tolist()  # 此处改密码
            self.diandata=df.index.tolist()#店名
            self.dc=df[4].tolist()#店名代码
            
        except(FileNotFoundError):
            print('请选择用户文件')
            data = self.getfile()
            df = pd.read_excel(data,header=None,index_col=3)
            self.username = df[0].tolist() # 此处改用户名
            self.password1 = df[1].tolist()  # 此处改密码
            self.diandata=df.index.tolist()#店名
            self.dc=df[4].tolist()#店名代码
        
    def dandian(self):
        try:
            data=r'信息表4.xlsx'
            df = pd.read_excel(data,header=None,index_col=3)
            name=askstring("请输入店名",prompt = "输入店名规则：宏发")
            list1=df.loc[name].tolist()
            user = list1[0] # 此处改用户名
            self.username=[user]
            password = list1[1] 
            self.password1=[password]# 此处改密码
            self.diandata=[name]#店名
            dcd=list1[3]#店名代码
            self.dc=[dcd]
        except(FileNotFoundError):
            print('请选择用户文件')
            data = self.getfile()
            df = pd.read_excel(data,header=None,index_col=3)
            name=askstring("请输入店名",prompt = "输入店名规则：宏发")
            list1=df.loc[name].tolist()
            user = list1[0] # 此处改用户名
            self.username=[user]
            password = list1[1] 
            self.password1=[password]# 此处改密码
            self.diandata=[name]#店名
            dcd=list1[3]#店名代码
            self.dc=[dcd]
        try:

            for n,p,dian in zip(self.username,self.password1,self.diandata):
                tmp = dw.Test(n,p)
                tmp.hex_pw()
                tmp.login(dian)
                time.sleep(2)
                tmp.zhihuan_down(self.startDate,self.endDate,dian,self.skcode,self.customer_name)#置换明细下载
                tmp.forcsv()
                tmp.hebing(dian)
                tmp.quit()
            print('所有下载进程完毕！！')
        except (AttributeError):
            msgbox.showerror('错误提示','请先输入日期和初始店面再点击按钮下载')

    def startdt(self):

        self.startDate=askstring("请输入开始日期",prompt = "开始日期：YYYY-MM-DD：")
        while self.fmdate(self.startDate)==False:
            print('开始日期格式有误,请重新输入')
            self.startDate=askstring("请输入开始日期",prompt = "开始日期：YYYY-MM-DD：")
        print('您输入的开始日期为：'+self.startDate)
#         self.endDate=askstring("请输入结束日期",prompt = "结束日期：YYYY-MM-DD：")  
    def enddt(self):

#         self.startDate=askstring("请输入开始日期",prompt = "开始日期：YYYY-MM-DD：")
#         while self.fmdate(self.startDate)==False:
#             print('开始日期格式有误,请重新输入')
#             self.startDate=askstring("请输入开始日期",prompt = "开始日期：YYYY-MM-DD：")
#         print('您输入的开始日期为：'+self.startDate)
        self.endDate=askstring("请输入结束日期",prompt = "结束日期：YYYY-MM-DD：")             
        while self.fmdate(self.endDate)==False:
            print('结束日期格式有误,请重新输入')
            self.endDate=askstring("请输入结束日期",prompt = "结束日期：YYYY-MM-DD：")
        print('您输入的结束日期为：'+self.endDate)
        if self.endDate<self.startDate:
            msgbox.showerror('错误提示', '开始日期不能大于结束日期哦，请重新点击输入日期')
#     def denglu(self):
    def zhihuan_down(self): 
        try:
            data=r'信息表4.xlsx'
            df = pd.read_excel(data,header=None,index_col=3)
            self.username = df[0].tolist() # 此处改用户名
            self.password1 = df[1].tolist()  # 此处改密码
            self.diandata=df.index.tolist()#店名
            self.dc=df[4].tolist()#店名代码
            
        except(FileNotFoundError):
            print('请选择用户文件')
            data = self.getfile()
            df = pd.read_excel(data,header=None,index_col=3)
            self.username = df[0].tolist() # 此处改用户名
            self.password1 = df[1].tolist()  # 此处改密码
            self.diandata=df.index.tolist()#店名
            self.dc=df[4].tolist()#店名代码
        print('请稍候，即将开始下载！')
        try:
            for n,p,dian in zip(self.username,self.password1,self.diandata):
                tmp = dw.Test(n,p)
                tmp.hex_pw()
                tmp.login(dian)
                time.sleep(2)
                tmp.zhihuan_down(self.startDate,self.endDate,dian,self.skcode,self.customer_name)#置换明细下载
                tmp.forcsv()
                tmp.hebing(dian)
                tmp.quit()
            print('所有下载进程完毕！！')
        except (AttributeError):
            msgbox.showerror('错误提示','请先输入日期和初始店面再点击按钮下载')
    def zhihuan_down1(self): 
        try:
            data=r'信息表4.xlsx'
            dadf = pd.read_excel(data,sheet_name=[0,1],header=None)
            df=dadf[0]
            df1=dadf[1]
            df.index=df[3]
            dian=askstring("请输入店名",prompt = "输入店名规则：宏发")
            list1=df.loc[dian].tolist()
            username = list1[0] # 此处改用户名
            password1 = list1[1]
            self.customer_name=df1[0].tolist()
        except(FileNotFoundError):
            print('请选择用户文件')
            data = self.getfile()
            dadf = pd.read_excel(data,sheet_name=[0,1],header=None)
            df=dadf[0]
            df1=dadf[1]
            df.index=df[3]
            dian=askstring("请输入店名",prompt = "输入店名规则：宏发")
            list1=df.loc[dian].tolist()
            username = list1[0] # 此处改用户名
            password1 = list1[1]
            self.customer_name=df1[0].tolist()

        print('请稍候，即将开始下载！')
        try:
            tmp = dw.Test(username,password1)
            tmp.hex_pw()
            tmp.login(dian)
            for name in self.customer_name:
                time.sleep(2)
                tmp.zhihuan_down(self.startDate,self.endDate,dian,self.skcode,name)#置换明细下载
                tmp.forcsv()
                tmp.hebing(name)

                print(name+'下载完成')
#                 print(self.startDate,self.endDate,dian,self.skcode,name,username,password1)
            tmp.quit()
            print('所有下载进程完毕！！')
#         except Exception as e:
#             print('出错了：', e)

        except (AttributeError):
            msgbox.showerror('错误提示','请先输入日期和初始店面再点击按钮下载')
#         self.kehuhebing()
        time.sleep(2)
        print('开始合并')
        hbfile=r'D:\pyoutdata\待上传置换增购\最终明细表\最终客户合并明细表.xlsx'
        filepath=r'D:\pyoutdata\待上传置换增购\最终明细表\置换'
        df1=pd.read_excel(hbfile,header=None)
        newdf=[]
        for name in self.customer_name:
            file=filepath+'\最终'+name+'合并明细表.xlsx'
            df3=pd.read_excel(file,header=None)
            newdf.append(df3)
            os.remove(file)
        newdf.append(df1)
        dkdf=pd.concat(newdf)
        dkdf.to_excel(filepath + '\最终'+ dian+'合并明细表1.xlsx',header=None,index=None)
        print('合并成功') 
    def mxsx(self):
        dianname=['宏发', '宏卓', '汇源', '宏诚', '宏哲', '宏之睿', '宏大']
        for name in dianname:
            daichab=r'D:\pyoutdata\待上传置换增购\最终明细表\置换'+'\最终'+name+'合并明细表.xlsx'#需匹配数据表位置
            mx=r'D:\jupython\DMS置换程序\长安2022年销量明细.xlsx'#待匹配数据汇总表位置
            mxdf=pd.read_excel(mx)#读取汇总明细表
            che=pd.read_excel(daichab,sheet_name=0,usecols=[2],header=None)#读取需匹配数据
        #     print(mxdf,che)
            daipipei=mxdf['车架号'].values#读取汇总表待匹配列数据
            daicha=che.values.tolist()#待匹配数据读取转为列表
            print(type(daipipei),daicha)
            j=[]
            k=[]
            for i in daicha:
        #         i=daicha[l][0]
        #         print(type(i[0]))
        #         w=np.where(daipipei==i[0])
                try:
                    w=np.where(daipipei==i[0])
        #             print(w[0][0])
                    j.append(w[0][0])#查找待匹配数据行号并记录列表
                except:
                    k.append(i[0])#未匹配数据保存列表
            print(j,k)
            newdf=mxdf.iloc[j]#根据匹配出的行号切片出匹配数据 
            weidf=pd.DataFrame({"未匹配数据":k})#未匹配数据转换为DF

            writer=pd.ExcelWriter(r'L:\2022置换增购\匹配后销售明细\\'+name+'筛选后数据.xlsx')#写入DF
            newdf.to_excel(writer,'筛选后数据',index=False)
            weidf.to_excel(writer,'未匹配数据',index=False)

            writer.save()
            print(name+'筛选完成！')
    def bpsx(self):
        zgdir = r'L:\2022置换增购\旧车信息表\置换\本品'
        mxdir = r'D:\pyoutdata\待上传置换增购\最终明细表\置换'
        zglist = os.listdir(zgdir) #列出文件夹下所有的目录与文件
        mxlist = os.listdir(mxdir) #列出文件夹下所有的目录与文
        for zg in range(0,len(zglist)):
            zgpath = os.path.join(zgdir,zglist[zg])
            mxpath = os.path.join(mxdir,mxlist[zg])
        #     print(zgpath)
        #     print(mxpath)
            name1=xlrd.open_workbook(mxpath)
            name1sheet = name1.sheet_by_name("Sheet1")
            namedata=name1sheet.col_values(9)
            chedata=name1sheet.col_values(20)
        #     print(namedata,chedata)
            workbook = openpyxl.load_workbook(zgpath)
            sheet = workbook["Sheet1"]

            demo_df=pd.read_excel(zgpath)

            # workbook = xlrd.open_workbook(file)
            # worksheet = workbook.sheet_by_name("Sheet1")
            # num_rows = worksheet.nrows
            # #查找数据所在行并写入
            newdata=[]
            z=0
            for n,j in zip(namedata,chedata):
                # print(type(i))
                # for worksheet_name in worksheets:
            #         worksheet = workbook.sheet_by_name(worksheet_name)
                for indexs in demo_df.index:
                    for i in range(len(demo_df.loc[indexs].values)):
                        if (demo_df.loc[indexs].values[i] == j):
                            # print(indexs+2)                
                            for row in sheet['G1':'G300']:
                                for cell in row:                        
                                    if j == cell.value:
                                        print(j)
                                        newdata.append(n)
                                        z=z+1                            
                                        sheet.cell(row=indexs+2, column=15, value=n)                        

            print(mxpath[32:-10]+'共查出以上'+str(z)+'个数据')
        #     print(str(list(set(namedata)-set(newdata)))+'未查到')
            workbook.save(zgpath)
#             print('查找成功')
        print('所有任务已完成')
    def tpsx(self):   
        zgdir = r'L:\2022置换增购\旧车信息表\置换\非本品'
        mxdir = r'D:\pyoutdata\待上传置换增购\最终明细表\置换'
        zglist = os.listdir(zgdir) #列出文件夹下所有的目录与文件
        mxlist = os.listdir(mxdir) #列出文件夹下所有的目录与文
        for zg in range(0,len(zglist)):
            zgpath = os.path.join(zgdir,zglist[zg])
            mxpath = os.path.join(mxdir,mxlist[zg])
        #     print(zgpath)
        #     print(mxpath)
            name1=xlrd.open_workbook(mxpath)
            name1sheet = name1.sheet_by_name("Sheet1")
            namedata=name1sheet.col_values(9)
            chedata=name1sheet.col_values(20)
        #     print(namedata,chedata)
            workbook = openpyxl.load_workbook(zgpath)
            sheet = workbook["Sheet1"]

            demo_df=pd.read_excel(zgpath)

            # workbook = xlrd.open_workbook(file)
            # worksheet = workbook.sheet_by_name("Sheet1")
            # num_rows = worksheet.nrows
            # #查找数据所在行并写入
            newdata=[]
            z=0
            for n,j in zip(namedata,chedata):
                # print(type(i))
                # for worksheet_name in worksheets:
            #         worksheet = workbook.sheet_by_name(worksheet_name)
                for indexs in demo_df.index:
                    for i in range(len(demo_df.loc[indexs].values)):
                        if (demo_df.loc[indexs].values[i] == j):
                            # print(indexs+2)                
                            for row in sheet['A1':'D101']:
                                for cell in row:                        
                                    if j == cell.value:
                                        print(j)
                                        newdata.append(n)
                                        z=z+1                            
                                        sheet.cell(row=indexs+2, column=9, value=n)                        

            print(mxpath[32:-10]+'共查出以上'+str(z)+'个数据')
        #     print(str(list(set(namedata)-set(newdata)))+'未查到')
            workbook.save(zgpath)
#             print('查找成功')
        print('所有任务已完成')
MyGUI() # 启动GUI

